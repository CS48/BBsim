import random, json, numpy, os.path, time, tkinter
from os import path
from openpyxl import Workbook, load_workbook

# some lists for later use
ages = [19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38]
positions = ["PG", "SG", "SF", "PF", "C", "PG", "SG", "SF", "PF", "C"]
first_name_data = ["Jacob", "Michael", "Matthew", "Joshua", "Christopher", "Nicholas", "Andrew", "Joseph", "Daniel",
                   "Tyler", "William", "Brandon", "Ryan", "John", "Zachary", "David", "Anthony", "James", "Justin",
                   "Alexander", "Jonathan", "Christian", "Austin", "Dylan", "Ethan", "Benjamin", "Noah", "Samuel",
                   "Robert", "Nathan", "Cameron", "Kevin", "Thomas", "Jose", "Hunter", "Jordan", "Kyle", "Caleb",
                   "Jason", "Logan", "Aaron", "Eric", "Brian", "Gabriel", "Adam", "Jack", "Isaiah", "Juan", "Luis",
                   "Connor", "Charles", "Elijah", "Isaac", "Steven", "Evan", "Jared", "Sean", "Timothy", "Luke",
                   "Cody", "Nathaniel", "Alex", "Seth", "Mason", "Richard", "Carlos", "Angel", "Patrick", "Devin",
                   "Bryan", "Cole"]
last_name_data = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Miller", "Davis", "Garcia", "Rodriguez", "Wilson",
                  "Martinez", "Anderson", "Taylor", "Thomas", "Hernandez", "Moore", "Martin", "Jackson", "Thompson",
                  "White", "Lopez", "Lee", "Gonzalez", "Harris", "Clark", "Lewis", "Robinson", "Walker", "Perez",
                  "Hall", "Young", "Allen", "Sanchez", "Wright", "King", "Scott", "Green", "Baker", "Adams", "Nelson",
                  "Hill", "Ramirez", "Campbell", "Mitchell", "Roberts"]
team_data = ["Atlanta", "Boston", "Charlotte", "Chicago", "Denver", "Detroit", "Houston", "Indiana", "Los Angeles",
             "Miami", "Minnesota"]

# for making a coin flip based on probability p. Such as an attack that has a 30% probability to work
def flip(p):
    return 'H' if random.random() < p else 'T'


def serialize_playerids_json(list):
    # opening the file in w mode should delete the data, so i don't have to worry about manually overwriting.
    file = 'player_ids.txt'
    filepath = os.path.abspath(os.getcwd())
    with open(os.path.join(filepath, file), 'w') as outfile:
        json.dump(list, outfile)

def serialize_gameids_json(list):
    # opening the file in w mode should delete the data, so i don't have to worry about manually overwriting.
    file = 'game_ids.txt'
    filepath = os.path.abspath(os.getcwd())
    with open(os.path.join(filepath, file), 'w') as outfile:
        json.dump(list, outfile)

def deserialize_playerids_json():
    file = 'player_ids.txt'
    filepath = os.path.abspath(os.getcwd())
    if path.exists(os.path.join(filepath, file)):
        with open(os.path.join(filepath, file)) as json_file:
            data = json.load(json_file)
        return list(data)
    else:
        print("No player id txt file.")

def deserialize_gameids_json():
    file = 'game_ids.txt'
    filepath = os.path.abspath(os.getcwd())
    if path.exists(os.path.join(filepath, file)):
        with open(os.path.join(filepath, file)) as json_file:
            data = json.load(json_file)
        return list(data)
    else:
        print("No game id txt file.")

def generate_id(list):
    if list is None:
        id = random.randint(1, 1000)
        return id
    else:
        pass

    while True:
        id = random.randint(1, 1000)
        if id in list:
            continue
        else:
            return id

def append_id(id, list):
    if list is None:
        list = [id]
        return list
    else:
        list.append(id)
        return list

def delete_id(id, list):
    if id in list:
        delete_this = list.index(id)
        del list[delete_this]
        return list
    else:
        print("Couldn't find id")

# the player class that we will use to build player objects from excel sheet data.
# it also contains some stats for tracking over the course of a game.
class Player:
    def __init__(self, arg_list):
        # player info

        self.player_id = arg_list[0]
        self.firstname = arg_list[1]
        self.lastname = arg_list[2]
        self.age = arg_list[3]
        self.position = arg_list[4]
        # stat tracking
        self.points = 0
        self.made_shots = 0
        self.shot_attempts = 0
        self.made_threes = 0
        self.three_attempts = 0
        self.assists = 0
        self.rebounds = 0
        self.blocks = 0
        self.steals = 0
        self.turnovers = 0
        self.time_played = 0
        self.time_restriction = 1680

        # attributes generate randomly based on player type and ovr is determined by a formula that is weighted
        # differently for each position

        self.inside_shot = arg_list[5]
        self.mid_shot = arg_list[6]
        self.three = arg_list[7]
        self.passing = arg_list[8]
        self.handling = arg_list[9]
        self.perimeter_d = arg_list[10]
        self.interior_d = arg_list[11]
        self.blocking = arg_list[12]
        self.stealing = arg_list[13]
        self.rebounding = arg_list[14]

    def update_restriction(self, n=int):
        self.time_restriction = n



# creates a new sheet in the Teams spreadsheet and generates some random players for that team by entering
# data into the cells. This spreadsheet will later be read to create player objects for simming
def create_team(team_name):
    # checks to see if the "Teams" file exists. Creates one if it doesn't
    if path.exists("Teams.xlsx"):
        pass
    else:
        workbook = Workbook()
        workbook.save(filename="Teams.xlsx")

    # makes sure that we are working in the Teams file
    filename = "Teams.xlsx"
    workbook = load_workbook(filename=filename)


    # the team name ends up as the name of the sheet within the Teams file
    team_sheet = workbook.create_sheet(team_name)

    # format sheet
    team_sheet["A1"] = "player_id"
    team_sheet["B1"] = "first_name"
    team_sheet["C1"] = "last_name"
    team_sheet["D1"] = "age"
    team_sheet["E1"] = "position"
    team_sheet["F1"] = "inside_shot"
    team_sheet["G1"] = "mid_shot"
    team_sheet["H1"] = "three"
    team_sheet["I1"] = "passing"
    team_sheet["J1"] = "handling"
    team_sheet["K1"] = "perimeter_d"
    team_sheet["L1"] = "interior_d"
    team_sheet["M1"] = "blocking"
    team_sheet["N1"] = "stealing"
    team_sheet["O1"] = "rebounding"

    # generates 5 players (one for every position) with random ratings
    for x in range(0, len(positions)):
        row = x+2

        # create_playerids_txt()
        print("hi")
        playerids = deserialize_playerids_json()
        new_id = generate_id(playerids)
        updated_list = append_id(new_id, playerids)
        serialize_playerids_json(updated_list)

        team_sheet["A%d" % row] = new_id
        team_sheet["B%d" % row] = random.choice(first_name_data)
        team_sheet["C%d" % row] = random.choice(last_name_data)
        team_sheet["D%d" % row] = random.choice(ages)
        team_sheet["E%d" % row] = positions[x]
        team_sheet["F%d" % row] = int(numpy.random.normal(65, 2, 1))
        team_sheet["G%d" % row] = int(numpy.random.normal(65, 2, 1))
        team_sheet["H%d" % row] = int(numpy.random.normal(65, 1, 1))
        team_sheet["I%d" % row] = int(numpy.random.normal(50, 15, 1))
        team_sheet["J%d" % row] = int(numpy.random.normal(50, 15, 1))
        team_sheet["K%d" % row] = int(numpy.random.normal(50, 15, 1))
        team_sheet["L%d" % row] = int(numpy.random.normal(50, 15, 1))
        team_sheet["M%d" % row] = int(numpy.random.normal(50, 15, 1))
        team_sheet["N%d" % row] = int(numpy.random.normal(50, 15, 1))
        team_sheet["O%d" % row] = int(numpy.random.normal(50, 15, 1))


    # saves the file, very important
    workbook.save(filename=filename)

    # Now we will to create a team records sheet to store the team's wins and losses
    if path.exists("Team_results.xlsx"):
        pass
    else:
        workbook = Workbook()
        workbook.save(filename="Team_results.xlsx")

    # I want to name the file "Team_results" and the sheet will be the name of the team.
    filename = "Team_results.xlsx"
    workbook = load_workbook(filename=filename)
    result_sheet = workbook.create_sheet(team_name)

    result_sheet["A1"] = "game_id"
    result_sheet["B1"] = "opponent"
    result_sheet["C1"] = "result"

    workbook.save(filename=filename)

    # feedback is good design ;)
    print("Team Successfully created\n")

# when we need to make player objects out of excel data, this is how we do it.
def load_team():
    # first we check that the Team file exists
    if path.exists("Teams.xlsx"):
        # then we make sure we are working in it
        filename = "Teams.xlsx"
        workbook = load_workbook(filename=filename)
        # we check that file to make sure that the team (given as an arg) is one of the sheets
        # within the file.
        while True:
            team_name = input("Which team would you like to load?:")
            if team_name in workbook.sheetnames:
                # If it exists, we select that sheet
                sheet = workbook[team_name]
                # this will be a list of player objects
                team = []
                # for each row in that sheet, which represents a player, we make a player object by
                # putting the data from that row into a list and feeding that list into Player() as
                # an arg. Then we put that player in the team list
                for x in range(2, sheet.max_row + 1):
                    for value in sheet.iter_rows(min_row=x, max_row=x, values_only=True):
                        arg_list = list(value)
                    player = Player(arg_list)
                    team.append(player)
                # return the team, so that we can do something with it.
                return team, team_name
            else:
                print("That team doesn't exist in the spreadsheet, try again")
                continue

# this is moreso a check than anything else. Just want to see that the player objects are
# successfully made from the excel sheet data
def print_team(team):
    for x in team:
        print(x.firstname, x.lastname)
        print(x.position, "\n")

        print("Inside Shot:", x.inside_shot)
        print("Mid Shot:", x.mid_shot)
        print("Three:", x.three )
        print("Passing:", x.passing)
        print("Handling:", x.handling)
        print("Perimeter D:", x.perimeter_d )
        print("Interior D:", x.interior_d)
        print("Blocking:", x.blocking )
        print("Stealing:", x.stealing)
        print("Rebounding:", x.rebounding, "\n")

# This will delete a sheet of data (effectively a team) from the Team file
def delete_team():
    # Check to make sure the Team file exists
    if path.exists("Teams.xlsx"):
        filename = "Teams.xlsx"
        workbook = load_workbook(filename=filename)

        # Show a list of the current sheets (Teams)
        print("Here are the current teams...")
        print(workbook.sheetnames, "\n")


        while True:
            # enter the name of a team to delete it
            team_name = input("Enter team name to delete:")
            if team_name in workbook.sheetnames:

                # this should update the id list to make sure that it is cleared of the deleted players
                sheet = workbook[team_name]
                ids_to_delete = []
                for x in range(2, sheet.max_row + 1):
                    ids_to_delete.append(sheet["A%d" % x].value)
                id_list = deserialize_playerids_json()
                print(ids_to_delete)
                print(id_list)
                for x in ids_to_delete:
                    id_list = delete_id(x, id_list)
                serialize_playerids_json(id_list)

                workbook.remove(workbook[team_name])
                break
            else:
                print("That team doesn't exist, try again")
                continue
        # save after it's done
        workbook.save(filename=filename)

        if path.exists("Team_results.xlsx"):
            filename = "Team_results.xlsx"
            workbook = load_workbook(filename=filename)
            workbook.remove(workbook[team_name])

            workbook.save(filename=filename)

        # feedback is good design
        print("%s has been deleted." % team_name)
    else:
        print("No teams currently.\n")

def show_teams():
    # for checking the sheets (Teams) in the Teams file
    if path.exists("Teams.xlsx"):
        filename = "Teams.xlsx"
        workbook = load_workbook(filename=filename)

        print(workbook.sheetnames, "\n")
    else:
        print("No teams exist currently.\n")


# So I'm considering a pass function. It's really not necessary in a text based sim, but i think it could be cool to
# be able to sim how the ball moves around the court. I'll have to determine how likely a pass it to happen. Maybe I'll
# create a variable called "defensive pressure" that determines how likely a pass is to happen. def_pressure could be
# determined by the defenders defense stats along with the ball handler's handling. It would also eventually need to
# consider the shot clock running down. This will be cool to code up.

# alright, so i'm just gonna try something here and see how it works
def Pass(ball_handler, previous, active_away, active_home, away_def_assign, home_def_assign, homepossession, shot_clock, time_remaining):
    pass_chance = 0
    passer = previous
    if homepossession:
        defender = away_def_assign[ball_handler]
    else:
        defender = home_def_assign[ball_handler]

    handler_shooting_avg = (ball_handler.three + ball_handler.mid_shot + ball_handler.inside_shot) / 3
    def_defense_avg = (defender.perimeter_d + defender.interior_d) / 2

    n = abs(handler_shooting_avg - def_defense_avg)

    if 0 < shot_clock <= 4:
        #print("The shot is clock is running down! Shot Clock: %d" % shot_clock)
        time_remaining = time_remaining - (24 - shot_clock)
        return ball_handler, passer, time_remaining
    elif shot_clock < 0:
        time_remaining = time_remaining - 24
        print("Shot clock violation! Turnover\n")
        return 1, None, time_remaining
    else:
        pass

    if handler_shooting_avg > def_defense_avg:
        if n >= 50:
            pass_chance = 0.20
        elif 25 <= n < 50:
            pass_chance = 0.40
        else:
            pass_chance = 0.60
    else:
        pass_chance = 0.75

    # noticed that pgs are taking wayyyyy too many shots
    if ball_handler.position == "PG":
        pass_chance = .85
    else:
        pass

    if flip(pass_chance) == 'H':
        if homepossession:
            while True:
                target = random.choice(list(active_home.values()))
                if target == ball_handler:
                    continue
                else:
                    break
            shot_clock = shot_clock - abs(int(numpy.random.normal(5, 2, 1)))
            #print("%s %s passes to %s %s. Shot Clock: %d" % (ball_handler.firstname, ball_handler.lastname, target.firstname
            #                                              ,target.lastname, shot_clock))
            return Pass(target, ball_handler, active_away, active_home, away_def_assign, home_def_assign, True, shot_clock, time_remaining)
        else:
            while True:
                target = random.choice(list(active_away.values()))
                if target == ball_handler:
                    continue
                else:
                    break
            shot_clock = shot_clock - abs(int(numpy.random.normal(5, 2, 1)))
            #print("%s %s passes to %s %s. Shot Clock: %d" % (ball_handler.firstname, ball_handler.lastname, target.firstname
            #,target.lastname, shot_clock))
            return Pass(target, ball_handler, active_away, active_home, away_def_assign, home_def_assign, False, shot_clock, time_remaining)
    else:
        time_remaining = time_remaining - (24 - shot_clock)
        return ball_handler, passer, time_remaining



        # select a team_mate besides oneself to pass to

def rebound(shooter, active_away, active_home, homepossession):
    away_rebounders = []
    home_rebounders = []
    for key in active_away:
        away_rebounders.append(active_away[key])
    for key in active_home:
        home_rebounders.append(active_home[key])


    # organize the away players from lowest rebounding to highest.
    for x in range(1, len(away_rebounders)):
        current_player = away_rebounders[x]
        current_value = away_rebounders[x].rebounding
        current_pos = x

        while current_pos > 0 and away_rebounders[current_pos - 1].rebounding > current_value:
            away_rebounders[current_pos] = away_rebounders[current_pos - 1]
            current_pos = current_pos - 1

        away_rebounders[current_pos] = current_player

    for x in range(1, len(home_rebounders)):
        current_player = home_rebounders[x]
        current_value = home_rebounders[x].rebounding
        current_pos = x

        while current_pos > 0 and home_rebounders[current_pos - 1].rebounding > current_value:
            home_rebounders[current_pos] = home_rebounders[current_pos - 1]
            current_pos = current_pos - 1

        home_rebounders[current_pos] = current_player


    if homepossession:
        if flip(0.25) == 'H':
            # offensive rebound
            rebounder = target = random.choices(home_rebounders, weights=(10, 10, 20, 25, 25), k=1)
            rebounder = rebounder[0]
            return rebounder, True
        else:
            #defensive rebound
            rebounder = target = random.choices(away_rebounders, weights=(10, 10, 20, 25, 25), k=1)
            rebounder = rebounder[0]
            return rebounder, False
    else:
        if flip(0.25) == 'H':
            # offensive rebound
            rebounder = target = random.choices(away_rebounders, weights=(10, 10, 20, 25, 25), k=1)
            rebounder = rebounder[0]
            return rebounder, False
        else:
            # defensive rebound
            rebounder = target = random.choices(home_rebounders, weights=(10, 10, 20, 25, 25), k=1)
            rebounder = rebounder[0]
            return rebounder, True

#outcomes of a shot attempt: made shot, missed shot, air ball out of bounds, blocked shot
def shoot(shooter, defender, passer):
    shot_score = 0
    def_score = 0

    # first i would like to determine the type of shot that they would most likely shoot based on rating.
    # For that, I would need to be able to determine how each of a player's three shooting ratings compare to each
    # other

    # 1 = close, 2 = mid, 3 = three
    shot_type = [1, 2, 3]

    # a very basic way of deciding whether a player would take more threes vs mid vs close shots
    if shooter.three >= 80 and shooter.mid_shot <= 80 and shooter.inside_shot <= 80:
        shot = random.choices(shot_type, weights=(25, 25, 50), k=1)
    elif shooter.three <= 80 and shooter.mid_shot >= 80 and shooter.inside_shot <= 80:
        shot = random.choices(shot_type, weights=(25, 50, 25), k=1)
    elif shooter.three <= 80 and shooter.mid_shot <= 80 and shooter.inside_shot >= 80:
        shot = random.choices(shot_type, weights=(50, 25, 25), k=1)
    else:
        shot = random.choices(shot_type, weights=(33, 33, 33), k=1)

    # I'm gonna do the same thing that i did with attacking in the vb sim
    # The shooter gets put into a qualitative category based on their relevant
    # shooting stat (0 = awful, 1 = bad, 2 = average, 3 = good, 4 = great). The same
    # is done for the defender. Then we determine the chance of the shooter
    # making the shot based off of that comparison.

    # stat track: add to shooter's shot attempts
    shooter.shot_attempts = shooter.shot_attempts + 1

    # determining shooter category
    # inside shot
    if shot[0] == 1:
        # output for console
        print("%s %s shoots from close range" % (shooter.firstname, shooter.lastname))

        if 0 <= shooter.inside_shot < 20:
            shot_score = shot_score + 0
        elif 20 <= shooter.inside_shot < 40:
            shot_score = shot_score + 1
        elif 40 <= shooter.inside_shot < 60:
            shot_score = shot_score + 2
        elif 60 <= shooter.inside_shot < 80:
            shot_score = shot_score + 3
        elif 80 <= shooter.inside_shot <= 100:
            shot_score = shot_score + 4
        else:
            print("shot_score error: stat not between 0 and 100")

        if 0 <= defender.interior_d < 20:
            def_score = def_score + 0
        elif 20 <= defender.interior_d < 40:
            def_score = def_score + 1
        elif 40 <= defender.interior_d < 60:
            def_score = def_score + 2
        elif 60 <= defender.interior_d < 80:
            def_score = def_score + 3
        elif 80 <= defender.interior_d <= 100:
            def_score = def_score + 4
        else:
            print("shot_score error: stat not between 0 and 100")

    # mid range shot
    elif shot[0] == 2:
        # output for console
        print("%s %s shoots from mid range" % (shooter.firstname, shooter.lastname))

        if 0 <= shooter.mid_shot < 20:
            shot_score = shot_score + 0
        elif 20 <= shooter.mid_shot < 40:
            shot_score = shot_score + 1
        elif 40 <= shooter.mid_shot < 60:
            shot_score = shot_score + 2
        elif 60 <= shooter.mid_shot < 80:
            shot_score = shot_score + 3
        elif 80 <= shooter.mid_shot <= 100:
            shot_score = shot_score + 4
        else:
            print("shot_score error: stat not between 0 and 100")

        if 0 <= ((defender.perimeter_d + defender.interior_d)/2) < 20:
            def_score = def_score + 0
        elif 20 <= ((defender.perimeter_d + defender.interior_d)/2) < 40:
            def_score = def_score + 1
        elif 40 <= ((defender.perimeter_d + defender.interior_d)/2) < 60:
            def_score = def_score + 2
        elif 60 <= ((defender.perimeter_d + defender.interior_d)/2) < 80:
            def_score = def_score + 3
        elif 80 <= ((defender.perimeter_d + defender.interior_d)/2) <= 100:
            def_score = def_score + 4
        else:
            print("shot_score error: stat not between 0 and 100")
    # three
    elif shot[0] == 3:
        # stat track: add to shooters three attempts
        shooter.three_attempts = shooter.three_attempts + 1
        # output for console
        print("%s %s shoots from three" % (shooter.firstname, shooter.lastname))

        if 0 <= shooter.three < 20:
            shot_score = shot_score + 0
        elif 20 <= shooter.three < 40:
            shot_score = shot_score + 1
        elif 40 <= shooter.three < 60:
            shot_score = shot_score + 2
        elif 60 <= shooter.three < 80:
            shot_score = shot_score + 3
        elif 80 <= shooter.three <= 100:
            shot_score = shot_score + 4
        else:
            print("shot_score error: stat not between 0 and 100")

        if 0 <= defender.perimeter_d < 20:
            def_score = def_score + 0
        elif 20 <= defender.perimeter_d < 40:
            def_score = def_score + 0
        elif 40 <= defender.perimeter_d < 60:
            def_score = def_score + 0
        elif 60 <= defender.perimeter_d < 80:
            def_score = def_score + 0
        elif 80 <= defender.perimeter_d <= 100:
            def_score = def_score + 0
        else:
            print("shot_score error: stat not between 0 and 100")

    else:
        print("Error in deciding type of shot")

    # get the success probability by calling the function and passing in the shot score and def score
    success_prob = shot_succ_prob(shot_score, def_score)



    # make a coin flip using the success probability to determine if the shot is made.
    if flip(success_prob) == 'H':
        # output for console
        print("It's good!")
        # i read somewhere that 50% of nba made shots are assisted
        if passer is None:
            pass
        else:
            if flip(.50) == 'H':
                print("%s %s with the assist." % (passer.firstname, passer.lastname))
                passer.assists = passer.assists + 1
            else:
                pass


        # the return will be used to determine how many points to give
        if shot[0] == 1 or shot[0] == 2:
            # stat track: add to made shots
            shooter.made_shots = shooter.made_shots + 1
            shooter.points = shooter.points + 2
            return 2
        elif shot[0] == 3:
            # stat track: add to made shot and made threes
            shooter.made_shots = shooter.made_shots + 1
            shooter.made_threes = shooter.made_threes + 1
            shooter.points = shooter.points + 3
            return 3
        else:
            print("error exiting shot flip")
    else:
        print("It's a missed shot")
        return 0


# convoluted bullshit. There is 100% a better method than this. Please find it.
# all it does is determine a success probability for a shot based on the qualitative category of the players' ratings.
def shot_succ_prob(shot_score, def_score):

    if shot_score == 0 and def_score == 0:
        success_prob = random.uniform(0.00, 0.40)
    elif shot_score == 0 and def_score == 1:
        success_prob = random.uniform(0.00, 0.30)
    elif shot_score == 0 and def_score == 2:
        success_prob = random.uniform(0.00, 0.20)
    elif shot_score == 0 and def_score == 3:
        success_prob = random.uniform(0.00, 0.10)
    elif shot_score == 0 and def_score == 4:
        success_prob = random.uniform(0.10, 0.60)
    elif shot_score == 1 and def_score == 0:
        success_prob = random.uniform(0.10, 0.60)
    elif shot_score == 1 and def_score == 1:
        success_prob = random.uniform(0.10, 0.50)
    elif shot_score == 1 and def_score == 2:
        success_prob = random.uniform(0.10, 0.40)
    elif shot_score == 1 and def_score == 3:
        success_prob = random.uniform(0.10, 0.30)
    elif shot_score == 1 and def_score == 4:
        success_prob = random.uniform(0.10, 0.20)
    elif shot_score == 2 and def_score == 0:
        success_prob = random.uniform(0.20, 0.70)
    elif shot_score == 2 and def_score == 1:
        success_prob = random.uniform(0.20, 0.60)
    elif shot_score == 2 and def_score == 2:
        success_prob = random.uniform(0.20, 0.50)
    elif shot_score == 2 and def_score == 3:
        success_prob = random.uniform(0.20, 0.40)
    elif shot_score == 2 and def_score == 4:
        success_prob = random.uniform(0.20, 0.30)
    elif shot_score == 3 and def_score == 0:
        success_prob = random.uniform(0.30, 0.80)
    elif shot_score == 3 and def_score == 1:
        success_prob = random.uniform(0.30, 0.70)
    elif shot_score == 3 and def_score == 2:
        success_prob = random.uniform(0.30, 0.60)
    elif shot_score == 3 and def_score == 3:
        success_prob = random.uniform(0.30, 0.50)
    elif shot_score == 3 and def_score == 4:
        success_prob = random.uniform(0.30, 0.40)
    elif shot_score == 4 and def_score == 0:
        success_prob = random.uniform(0.40, 0.90)
    elif shot_score == 4 and def_score == 1:
        success_prob = random.uniform(0.40, 0.80)
    elif shot_score == 4 and def_score == 2:
        success_prob = random.uniform(0.40, 0.70)
    elif shot_score == 4 and def_score == 3:
        success_prob = random.uniform(0.40, 0.60)
    elif shot_score == 4 and def_score == 4:
        success_prob = random.uniform(0.40, 0.50)
    else:
        print("problem with success prob")

    return success_prob


def play_quarter(awayteam, hometeam, active_away, active_home, away_def_assign, home_def_assign, awaypoints, homepoints, home_possession,
                 time_remaining, passed_time, offensive_rebound):
    # time.sleep(1)

    # I'm gonna try some minutes rotation stuff out. For that we need to keep track of how long a player has been
    # in the game.

    for x in range(1, 6):
        active_away[x].time_played = active_away[x].time_played + passed_time

        if active_away[x].time_played > active_away[x].time_restriction:
            for y in awayteam:
                if y.position == active_away[x].position:
                    sub_in = y

                    if sub_in.player_id == active_away[x].player_id:
                        pass
                    else:
                        sub_out = active_away[x]
                        active_away[x] = sub_in
                        print("%s %s subbing in for %s %s" % (sub_in.firstname, sub_in.lastname, sub_out.firstname,
                                                              sub_out.lastname))
                        for key in away_def_assign:
                            if away_def_assign[key] == sub_out:
                                away_def_assign[key] = sub_in
                                break
                            else:
                                pass

                        for key in home_def_assign:
                            if key == sub_out:
                                home_def_assign[sub_in] = home_def_assign[sub_out]
                                del home_def_assign[sub_out]
                                break
                else:
                    pass
        else:
            pass

    for x in range(1, 6):
        active_home[x].time_played = active_home[x].time_played + passed_time

        if active_home[x].time_played > active_home[x].time_restriction:
            for y in hometeam:
                if y.position == active_home[x].position:
                    sub_in = y

                    if sub_in.player_id == active_home[x].player_id:
                        pass
                    else:
                        sub_out = active_home[x]
                        active_home[x] = sub_in
                        print("%s %s subbing in for %s %s" % (sub_in.firstname, sub_in.lastname, sub_out.firstname,
                                                              sub_out.lastname))
                        for key in home_def_assign:
                            if home_def_assign[key] == sub_out:
                                home_def_assign[key] = sub_in
                                break
                            else:
                                pass

                        for key in away_def_assign:
                            if key == sub_out:
                                away_def_assign[sub_in] = away_def_assign[sub_out]
                                del away_def_assign[sub_out]
                                break
                else:
                    pass
        else:
            pass



    time_remaining = time_remaining - abs(int(numpy.random.normal(4, 1, 1)))

    if time_remaining > 0:
        # convert the seconds to minutes format
        mins, sec = divmod(time_remaining, 60)
        print("Time:  %d:%d" % (mins, sec))

        if home_possession:
            print("Home possession")
            # ball handler is pg unless offensive rebound
            if offensive_rebound[0] == True:
                ball_handler = offensive_rebound[1]
                shotclock = 14
            else:
                active_players = list(active_home.values())
                ball_handler = active_players[0]
                shotclock = 20


            # if the time left is less than 25, then turn off the shot clock
            if time_remaining < 25:
                # go into pass function with shot clock == time remaining
                pass_result = Pass(ball_handler, None, active_away, active_home, away_def_assign, home_def_assign, home_possession, time_remaining, time_remaining)
            else:
                # go into pass function with a 20 second shot clock (20 to simulate moving the ball over half court)
                pass_result = Pass(ball_handler, None, active_away, active_home, away_def_assign, home_def_assign,
                                   home_possession, shotclock, time_remaining)

            # keep up with how much time has pass in a possession to update minutes for the players

            old_time = time_remaining
            time_remaining = pass_result[2]
            time_passed = abs(old_time - time_remaining)
            # a print statement for error checking. Can be deleted
            print("%d seconds have passed." % time_passed)
            if pass_result[0] == 1:
                # if there was a shot clock violation, we run the play quarter function again with the other team in
                # possesion
                return play_quarter(awayteam, hometeam, active_away, active_home, away_def_assign, home_def_assign, awaypoints, homepoints,
                                    False, time_remaining, time_passed, [False])
            else:
                pass

            # get ready to shoot the ball

            # look at the away defensive assignments to select the right defender
            defender = away_def_assign[pass_result[0]]
            # call the shot function (args = shooter, defender, passer)
            shot_result = shoot(pass_result[0], defender, pass_result[1])

            # assign points based on results
            if shot_result == 2:
                homepoints = homepoints + 2
            elif shot_result == 3:
                homepoints = homepoints + 3
            elif shot_result == 0:
                rebound_result = rebound(pass_result[0], active_away, active_home, home_possession)
                rebounder = rebound_result[0]
                # stat track: add a rebound
                rebounder.rebounds = rebounder.rebounds + 1
                home_possession = rebound_result[1]

                if home_possession:
                    print("%s %s grabs the offensive rebound.\n" % (rebounder.firstname, rebounder.lastname))
                    return play_quarter(awayteam, hometeam, active_away, active_home, away_def_assign, home_def_assign, awaypoints, homepoints,
                                        home_possession, time_remaining, time_passed, [True, rebounder])
                else:
                    print("%s %s grabs the rebound.\n" % (rebounder.firstname, rebounder.lastname))
                    return play_quarter(awayteam, hometeam, active_away, active_home, away_def_assign, home_def_assign, awaypoints, homepoints,
                                 home_possession, time_remaining, time_passed, [False])
            else:
                print("problem with points")

            print("\nScore:", awaypoints, homepoints, "\n")

            # recursively call the function again for the next possession where the other team has the ball after a
            # made shot
            return play_quarter(awayteam, hometeam, active_away, active_home, away_def_assign, home_def_assign, awaypoints, homepoints,
                                False, time_remaining, time_passed, [False])

        # Away possession
        else:
            print("Away possession")
            # ball handler is pg unless offensive rebound
            if offensive_rebound[0] == True:
                ball_handler = offensive_rebound[1]
                shotclock = 14
            else:
                active_players = list(active_away.values())
                ball_handler = active_players[0]
                shotclock = 24

            # if time_remaining is 24 or less, then shot clock == time remaining, else pass in 20 for shot clock
            # (20 to simulate moving the ball over half court)
            if time_remaining < 25:
                pass_result = Pass(ball_handler, None, active_away, active_home, away_def_assign, home_def_assign,
                                   home_possession,
                                   time_remaining, time_remaining)
            else:
                pass_result = Pass(ball_handler, None, active_away, active_home, away_def_assign, home_def_assign, home_possession,
                               shotclock, time_remaining)

            # keep track of possession time to update minutes for players
            old_time = time_remaining
            time_remaining = pass_result[2]
            time_passed = abs(old_time - time_remaining)
            # error checking statement. can be deleted
            print("%d seconds have passed." % time_passed)

            # if shot clock violation, other team gets ball.
            if pass_result[0] == 1:
                return play_quarter(awayteam, hometeam, active_away, active_home, away_def_assign, home_def_assign, awaypoints, homepoints,
                                    True, time_remaining, time_passed, [False])
            else:
                pass

            #get ready to shoot

            #get the right defender from the defensive assignment dict
            defender = home_def_assign[pass_result[0]]
            # the function to shoot the ball (args = shooter, defender, passer)
            shot_result = shoot(pass_result[0], defender, pass_result[1])

            if shot_result == 2:
                awaypoints = awaypoints + 2
            elif shot_result == 3:
                awaypoints = awaypoints + 3
            elif shot_result == 0:
                rebound_result = rebound(pass_result[0], active_away, active_home, home_possession)
                rebounder = rebound_result[0]
                # stat track: add a rebound
                rebounder.rebounds = rebounder.rebounds + 1
                home_possession = rebound_result[1]

                if home_possession:
                    print("%s %s grabs the rebound.\n" % (rebounder.firstname, rebounder.lastname))
                    return play_quarter(awayteam, hometeam, active_away, active_home, away_def_assign, home_def_assign, awaypoints,
                                        homepoints,
                                        home_possession, time_remaining, time_passed, [False])
                else:
                    print("%s %s grabs the offensive rebound.\n" % (rebounder.firstname, rebounder.lastname))
                    return play_quarter(awayteam, hometeam, active_away, active_home, away_def_assign, home_def_assign, awaypoints, homepoints,
                                 home_possession, time_remaining, time_passed, [True, rebounder])
            else:
                print("problem with points")

            print("\nScore:", awaypoints, homepoints, "\n")

            return play_quarter(awayteam, hometeam, active_away, active_home, away_def_assign, home_def_assign, awaypoints, homepoints,
                                True, time_remaining, time_passed, [False])

    else:
        return awaypoints, homepoints


# Work in progress. The actually act of simulating a game. Gonna break it down by quarters I guess.
def play(slow=None):
    # make sure there is a teams file before we do anything
    if path.exists("Teams.xlsx"):
        pass
    else:
        print("There are no teams currently.")
        return 0


    # work in the teams file
    filename = "Teams.xlsx"
    workbook = load_workbook(filename=filename)
    # let the user select the two teams that are gonna play
    print("Here are the teams that are available:\n")
    show_teams()

    print("Who is the away team?")
    result = load_team()
    away_team = result[0]
    away_name = result[1]

    print("Who is the home team?")
    result = load_team()
    home_team = result[0]
    home_name = result[1]


    # I'm gonna use a dict to assign players from the team to the 5 active spots on the floor. This will be useful
    # later if you have teams of more than 5 players
    active_away = {1: None, 2: None, 3: None, 4: None, 5: None}
    active_home = {1: None, 2: None, 3: None, 4: None, 5: None}

    # assigning to active dicts
    for x in range(1, 6):
        active_away[x] = away_team[x - 1]
        active_home[x] = home_team[x - 1]

    # printing out a list to check
    print("Away team active players\n")
    for x in range(1,6):
        print("%s %s" % (active_away[x].firstname, active_away[x].lastname))
    print("\nHome team active players\n")
    for x in range(1,6):
        print("%s %s" % (active_home[x].firstname, active_home[x].lastname))


    # So this is just an idea, but i thought it would be cool to be able to assign defenders. So in order to do that,
    # you need to know what defender to reference whenever a certain player has the ball. I can do this by creating a
    # dict where the active players on one team are the keys and the active players on the other team are the values.
    # I'll need one of these for both teams because their assignments may not mirror each other.

    away_def_assign = {active_home[1]: active_away[1], active_home[2]: active_away[2], active_home[3]: active_away[3]
        , active_home[4]: active_away[4], active_home[5]: active_away[5]}
    home_def_assign = {active_away[1]: active_home[1], active_away[2]: active_home[2], active_away[3]: active_home[3]
        , active_away[4]: active_home[4], active_away[5]: active_home[5]}

    print("Away Defensive Assignments\n")
    for x in away_def_assign:
        print("%s %s : %s %s\n" % (x.firstname, x.lastname, away_def_assign[x].firstname, away_def_assign[x].lastname))
    print("Home Defensive Assignment\n")
    for x in home_def_assign:
        print("%s %s : %s %s\n" % (x.firstname, x.lastname, home_def_assign[x].firstname, home_def_assign[x].lastname))

    #1st quarter
    print("\nBeginning of 1st Quarter\n")
    quarter_result = play_quarter(away_team, home_team, active_away, active_home, away_def_assign, home_def_assign, 0, 0, False, 720, 0, [False])
    print("\nEnd of 1st Quarter")
    print("Score: %d %d\n" % (quarter_result[0], quarter_result[1]))
    #2nd quarter
    #time.sleep(2)

    print("\nBeginning of 2nd Quarter\n")
    quarter_result = play_quarter(away_team, home_team, active_away, active_home, away_def_assign, home_def_assign, quarter_result[0],
                                  quarter_result[1], False, 720, 0,[False])
    print("\nEnd of 2nd Quarter")
    print("Score: %d %d\n" % (quarter_result[0], quarter_result[1]))

    #3rd quarter
    #time.sleep(2)
    print("\nBeginning of 3rd Quarter\n")
    quarter_result = play_quarter(away_team, home_team, active_away, active_home, away_def_assign, home_def_assign, quarter_result[0],
                                  quarter_result[1], False, 720, 0, [False])
    print("\nEnd of 3rd Quarter")
    print("Score: %d %d\n" % (quarter_result[0], quarter_result[1]))

    #4th quarter
    #time.sleep(2)
    print("\nBeginning of 4th Quarter\n")
    quarter_result = play_quarter(away_team, home_team, active_away, active_home, away_def_assign, home_def_assign, quarter_result[0],
                                  quarter_result[1], False, 720, 0, [False])
    print("\nEnd of 4th Quarter")
    print("Score: %d %d\n" % (quarter_result[0], quarter_result[1]))

    while True:
        #time.sleep(2)
        overtime_no = 1
        if quarter_result[0] == quarter_result[1]:
            print("\nBeginning of OT%d\n" % overtime_no)
            quarter_result = play_quarter(away_team, home_team, active_away, active_home, away_def_assign, home_def_assign, quarter_result[0],
                                          quarter_result[1], False, 720, 0, [False])
            continue
        else:
            break


    if quarter_result[0] > quarter_result[1]:
        print("\n%s wins\n" %away_name)
    else:
        print("\n%s wins\n" %home_name)

    # log game result in the team_results file
    # first we need to give the game an id
    game_ids = deserialize_gameids_json()
    new_gameid = generate_id(game_ids)
    updated_list = append_id(new_gameid, game_ids)
    serialize_gameids_json(updated_list)

    # updating team results sheet for away team
    filename = "Team_results.xlsx"
    workbook = load_workbook(filename=filename)
    awayteam_sheet = workbook[away_name]
    x = awayteam_sheet.max_row + 1
    awayteam_sheet["A%d" % x] = new_gameid
    awayteam_sheet["B%d" % x] = home_name
    if quarter_result[0] > quarter_result[1]:
        awayteam_sheet["C%d" % x] = "W"
    else:
        awayteam_sheet["C%d" % x] = "L"
    workbook.save(filename=filename)

    # updating team results sheet for home team
    hometeam_sheet = workbook[home_name]
    x = hometeam_sheet.max_row + 1
    hometeam_sheet["A%d" % x] = new_gameid
    hometeam_sheet["B%d" % x] = away_name
    if quarter_result[0] > quarter_result[1]:
        hometeam_sheet["C%d" % x] = "L"
    else:
        hometeam_sheet["C%d" % x] = "W"
    workbook.save(filename=filename)

    print("%s stats\n" % away_name)
    # let's try printing some stats
    for x in away_team:
        print("%s %s's stats" % (x.firstname, x.lastname))
        print("Points: %d" % x.points)
        print("FG: %d/%d" % (x.made_shots, x.shot_attempts))
        print("3 FG: %d/%d" % (x.made_threes, x.three_attempts))
        print("Assist: %d" % x.assists)
        print("Rebounds: %d" % x.rebounds)
        print("Minutes Played: %d\n" % int(x.time_played / 60))


    print("\n")
    print("%s stats\n" % home_name)

    for x in home_team:
        print("%s %s's stats" % (x.firstname, x.lastname))
        print("Points: %d" % x.points)
        print("FG: %d/%d" % (x.made_shots, x.shot_attempts))
        print("3 FG: %d/%d" % (x.made_threes, x.three_attempts))
        print("Assists: %d" % x.assists)
        print("Rebounds: %d" % x.rebounds)
        print("Minutes Played: %d\n" % int(x.time_played/60))

    print("\n")

    # This section for saving stats of all players who played in game xlsx
    if not path.exists('Game_stats.xlsx'):
        workbook = Workbook()
        workbook.save(filename="Game_stats.xlsx")
    else:
        pass
    # makes sure that we are working in the Game_stats file
    filename = "Game_stats.xlsx"
    workbook = load_workbook(filename=filename)
    gamestats_sheet = workbook.create_sheet(str(new_gameid))

    gamestats_sheet["A1"] = "player_id"
    gamestats_sheet["B1"] = "first_name"
    gamestats_sheet["C1"] = "last_name"
    gamestats_sheet["D1"] = "team"
    gamestats_sheet["E1"] = "minutes"
    gamestats_sheet["F1"] = "points"
    gamestats_sheet["G1"] = "FG"
    gamestats_sheet["H1"] = "3FG"
    gamestats_sheet["I1"] = "assists"
    gamestats_sheet["J1"] = "rebounds"

    row = 1
    for player in away_team:
        row = row + 1
        gamestats_sheet["A%d" % row] = player.player_id
        gamestats_sheet["B%d" % row] = player.firstname
        gamestats_sheet["C%d" % row] = player.lastname
        gamestats_sheet["D%d" % row] = away_name
        gamestats_sheet["E%d" % row] = int(player.time_played / 60)
        gamestats_sheet["F%d" % row] = player.points
        gamestats_sheet["G%d" % row] = "%d/%d" % (player.made_shots, player.shot_attempts)
        gamestats_sheet["H%d" % row] = "%d/%d" % (player.made_threes, player.three_attempts)
        gamestats_sheet["I%d" % row] = player.assists
        gamestats_sheet["J%d" % row] = player.rebounds

    for player in home_team:
        row = row + 1
        gamestats_sheet["A%d" % row] = player.player_id
        gamestats_sheet["B%d" % row] = player.firstname
        gamestats_sheet["C%d" % row] = player.lastname
        gamestats_sheet["D%d" % row] = home_name
        gamestats_sheet["E%d" % row] = int(player.time_played / 60)
        gamestats_sheet["F%d" % row] = player.points
        gamestats_sheet["G%d" % row] = "%d/%d" % (player.made_shots, player.shot_attempts)
        gamestats_sheet["H%d" % row] = "%d/%d" % (player.made_threes, player.three_attempts)
        gamestats_sheet["I%d" % row] = player.assists
        gamestats_sheet["J%d" % row] = player.rebounds

    workbook.save(filename=filename)



def create_save():

    x = input("Save Name?: ")

    current_directory = os.getcwd()
    final_directory = os.path.join(current_directory, r'%s' % x)
    if not os.path.exists(final_directory):
        os.makedirs(final_directory)
        print("Directory Created: %s" % x)

    os.chdir(final_directory)
    print("Entered Directory: %s\n" % x)

    for y in team_data:
        create_team(y)

def load_save():
    current_directory = os.getcwd()
    for file in os.listdir(current_directory):
        d = os.path.join(current_directory, file)
        if os.path.isdir(d):
            print(d)

    x = input("Choose save: ")
    final_directory = os.path.join(current_directory, r'%s' % x)
    os.chdir(final_directory)
    print("Entered Directory: %s\n" % x)

def main():
    # This is effectively a console based menu that keeps running until you exit it
    # the menu options are selected by entering a number that corresponds with that option.
    # Run main() and give it a try.

    print("Menu:\n"
          "1. Create a Save\n"
          "2. Load a Save")

    selection = input("Input a number:")
    if selection == "1":
        create_save()
    elif selection == "2":
        load_save()

    while True:
        print("Menu:\n"
              "1. Generate a Team\n"
              "2. Delete a Team\n"
              "3. See Current Teams\n"
              "4. Load Teams\n"
              "5. Play a game\n"
              "6. Exit")

        selection = input("Input a number:")
        if selection == "1":
            print("Creating a team for you...")
            create_team()
            continue
        elif selection == "2":
            delete_team()
            continue
        elif selection == "3":
            show_teams()
            continue
        elif selection == "4":
            print("Current Teams:")
            show_teams()
            team = load_team()
            print_team(team[0])
            print("\n")
            continue
        elif selection == "5":
            play()
            continue
        elif selection == "6":
            s = input("Are you sure (y/n):")
            if s == "y":
                break
            else:
                continue
        else:
            print("Invalid input, try again.\n")
            continue


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
